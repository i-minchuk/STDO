from __future__ import annotations
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from typing import Optional, List, Dict
from core.auth import get_current_user, require_role
from core.service_locator import get_locator
from models.user import User
import json

router = APIRouter(prefix="/api/import", tags=["import"])


class ColumnMapping(BaseModel):
    source_column: str
    target_field: str


class ImportConfig(BaseModel):
    project_id: int
    sheet_name: Optional[str] = None
    header_row: int = 0
    mappings: List[ColumnMapping] = []
    custom_columns: List[str] = []  # дополнительные колонки автора


@router.post("/excel/preview")
async def preview_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Загрузка Excel файла — возвращает список листов, колонок и первые 5 строк для превью."""
    try:
        import openpyxl
        from io import BytesIO
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(c) if c else f"Колонка_{i+1}" for i, c in enumerate(rows[0])]
            preview_rows = []
            for row in rows[1:6]:
                preview_rows.append({h: (str(v) if v is not None else "") for h, v in zip(headers, row)})
            sheets.append({
                "name": sheet_name,
                "columns": headers,
                "total_rows": len(rows) - 1,
                "preview": preview_rows,
            })
        wb.close()
        return {"filename": file.filename, "sheets": sheets}
    except Exception as e:
        raise HTTPException(400, f"Ошибка чтения файла: {str(e)}")


# Целевые поля системы
TARGET_FIELDS = [
    {"field": "document_code", "label": "Код документа", "required": True},
    {"field": "title", "label": "Наименование", "required": True},
    {"field": "doc_type", "label": "Тип документа", "required": False},
    {"field": "discipline", "label": "Дисциплина", "required": False},
    {"field": "engineer", "label": "Исполнитель", "required": False},
    {"field": "reviewer", "label": "Проверяющий", "required": False},
    {"field": "planned_start", "label": "Дата начала (план)", "required": False},
    {"field": "planned_finish", "label": "Дата окончания (план)", "required": False},
    {"field": "planned_hours", "label": "Плановые часы", "required": False},
    {"field": "ifr_date", "label": "Дата IFR", "required": False},
    {"field": "ifa_date", "label": "Дата IFA", "required": False},
    {"field": "ifc_date", "label": "Дата IFC", "required": False},
    {"field": "status", "label": "Статус", "required": False},
    {"field": "revision", "label": "Ревизия", "required": False},
    {"field": "notes", "label": "Примечания", "required": False},
]


@router.get("/target-fields")
def get_target_fields(current_user: User = Depends(get_current_user)):
    """Возвращает список полей системы, к которым можно привязать колонки из Excel."""
    return TARGET_FIELDS


@router.post("/excel/execute")
async def execute_import(
    config: str = "",  # JSON ImportConfig
    file: UploadFile = File(...),
    current_user: User = Depends(require_role("admin", "manager")),
):
    """Импорт данных из Excel в проект по настроенному маппингу."""
    try:
        import openpyxl
        from io import BytesIO

        cfg = ImportConfig(**json.loads(config)) if config else ImportConfig(project_id=0)
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb[cfg.sheet_name] if cfg.sheet_name else wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(c) if c else "" for c in rows[cfg.header_row]]

        mapping_dict = {m.source_column: m.target_field for m in cfg.mappings}
        imported = []
        errors = []

        for i, row in enumerate(rows[cfg.header_row + 1:], start=cfg.header_row + 2):
            row_data = {}
            for col_name, value in zip(headers, row):
                if col_name in mapping_dict:
                    row_data[mapping_dict[col_name]] = str(value) if value is not None else ""
            for custom_col in cfg.custom_columns:
                if custom_col in headers:
                    idx = headers.index(custom_col)
                    row_data[f"custom_{custom_col}"] = str(row[idx]) if row[idx] is not None else ""

            if not row_data.get("document_code"):
                errors.append({"row": i, "error": "Отсутствует код документа"})
                continue
            imported.append(row_data)

        wb.close()
        return {
            "imported_count": len(imported),
            "errors_count": len(errors),
            "errors": errors[:10],
            "preview": imported[:5],
        }
    except Exception as e:
        raise HTTPException(400, f"Ошибка импорта: {str(e)}")
