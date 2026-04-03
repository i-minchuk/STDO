from __future__ import annotations
from typing import Literal
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

ReportType = Literal["project", "projects", "employees", "labor"]

REPORT_HEADERS: dict[ReportType, list[str]] = {
    "labor": [
        "Сотрудник", "Проект", "Дисциплина",
        "Плановые ч/ч", "Факт. ч/ч", "% выполнения",
        "Ставка ₽/ч", "Плановая стоимость ₽", "Факт. стоимость ₽",
        "Отклонение ₽", "Накладные (25%) ₽", "ИТОГО ₽", "Рентабельность",
    ],
    "project": [
        "Документ", "Код", "Дисциплина", "Ответственный", "Статус",
        "Ревизия", "Плановые ч/ч", "Факт. ч/ч",
        "Срок сдачи (план)", "Факт. сдача", "Просрочка (дни)",
    ],
    "projects": [
        "Проект", "Код", "Заказчик", "Статус", "Плановые ч/ч",
        "Факт. ч/ч", "% выполнения", "Дата начала", "Дата окончания (план)",
    ],
    "employees": [
        "Сотрудник", "Роль", "Дисциплина", "Проект",
        "Плановые ч/ч", "Факт. ч/ч", "Задач завершено", "Очки (геймиф.)",
    ],
}


def generate_excel_report(report_type: ReportType, data: list[dict],
                           title: str) -> bytes:
    assert OPENPYXL_OK, "openpyxl не установлен. Добавь в requirements.txt"

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Стили шапки
    h_fill = PatternFill("solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", size=10)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = REPORT_HEADERS[report_type]

    # Строка-заголовок
    ws.row_dimensions[1].height = 36
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = 20

    # Данные
    money_cols: set[int] = set()
    pct_cols: set[int] = set()
    if report_type == "labor":
        money_cols = {8, 9, 10, 11, 12}
        pct_cols = {13}

    alt_fill = PatternFill("solid", fgColor="F2F7FC")

    for ri, row_data in enumerate(data, 2):
        row_fill = alt_fill if ri % 2 == 0 else None
        for ci, key in enumerate(row_data.keys(), 1):
            val = row_data[key]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill
            if ci in money_cols:
                cell.number_format = '# ##0.00 [$₽-419]'
            if ci in pct_cols:
                cell.number_format = '0.0"%"'
                if isinstance(val, (int, float)) and val < 0:
                    cell.fill = PatternFill("solid", fgColor="FFD0D0")

    # Итоговая строка для labor
    if report_type == "labor" and data:
        total_row = len(data) + 2
        ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
        for ci in money_cols:
            col_letter = get_column_letter(ci)
            ws.cell(row=total_row, column=ci,
                    value=f"=SUM({col_letter}2:{col_letter}{total_row-1})"
                    ).font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="D6E4F0")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()